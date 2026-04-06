from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
import asyncpg, aiofiles, uuid, asyncio, io
import os, logging, bcrypt, jwt, secrets, string, json
import httpx
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

DATABASE_URL = os.getenv('DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/lionjalyuzidb')
JWT_SECRET = os.getenv('JWT_SECRET', 'lionjalyuzi-super-secret-key-2024')

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI()
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
api_router = APIRouter(prefix="/api")
JWT_ALGORITHM = "HS256"

# ─── IN-MEMORY CACHE (DB tezlashtirish) ───
import time as _time

class MemoryCache:
    """Oddiy xotiradagi kesh — DB so'rovlarni 3-5x tezlashtiradi"""
    def __init__(self):
        self._store: dict = {}
        self._ttl: dict = {}

    def get(self, key: str):
        if key in self._store and _time.time() < self._ttl.get(key, 0):
            return self._store[key]
        self._store.pop(key, None)
        self._ttl.pop(key, None)
        return None

    def set(self, key: str, value, ttl_seconds: int = 30):
        self._store[key] = value
        self._ttl[key] = _time.time() + ttl_seconds

    def invalidate(self, *prefixes):
        """Berilgan prefikslar bilan boshlanadigan barcha keshlarni o'chiradi"""
        keys_to_del = [k for k in self._store if any(k.startswith(p) for p in prefixes)]
        for k in keys_to_del:
            self._store.pop(k, None)
            self._ttl.pop(k, None)

    def clear(self):
        self._store.clear()
        self._ttl.clear()

cache = MemoryCache()

# ─── DB Pool ───
pool: asyncpg.Pool = None

async def get_pool() -> asyncpg.Pool:
    global pool
    if pool is None:
        pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)
    return pool

# ─── Helpers ───
def get_jwt_secret(): return os.environ["JWT_SECRET"]
def hash_password(pw: str) -> str: return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
def verify_password(plain: str, hashed: str) -> bool: return bcrypt.checkpw(plain.encode(), hashed.encode())

def create_access_token(uid: str, email: str, role: str) -> str:
    return jwt.encode({"sub": uid, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def generate_order_code():
    return ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(8))

def row_to_dict(row):
    if row is None:
        return None
    return dict(row)

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "): raise HTTPException(401, "Not authenticated")
    try:
        p = jwt.decode(auth[7:], get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        try:
            user_id = int(p["sub"])
        except (ValueError, TypeError):
            raise HTTPException(401, "Invalid token format - please login again")
        db = await get_pool()
        row = await db.fetchrow("SELECT * FROM users WHERE id = $1", user_id)
        if not row: raise HTTPException(401, "User not found")
        user = row_to_dict(row)
        user["id"] = str(user["id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError: raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError: raise HTTPException(401, "Invalid token")
    except HTTPException: raise
    except Exception as e:
        logger.error(f"Auth error: {e}")
        raise HTTPException(401, "Authentication failed")

async def require_admin(request: Request) -> dict:
    u = await get_current_user(request)
    if u.get("role") != "admin": raise HTTPException(403, "Admin only")
    return u

async def require_worker(request: Request) -> dict:
    u = await get_current_user(request)
    if u.get("role") != "worker": raise HTTPException(403, "Worker only")
    return u

# ─── Pydantic Models ───
class LoginReq(BaseModel): email: str; password: str
class DealerCreate(BaseModel): name: str; email: str; password: str; phone: str = ""; address: str = ""; credit_limit: float = 0
class DealerUpdate(BaseModel): name: Optional[str] = None; phone: Optional[str] = None; address: Optional[str] = None; credit_limit: Optional[float] = None
class WorkerCreate(BaseModel): name: str; email: str; password: str; phone: str = ""; specialty: str = ""
class MaterialCreate(BaseModel): name: str; category: str = ""; category_id: Optional[int] = None; price_per_sqm: float; stock_quantity: float; unit: str = "kv.m"; description: str = ""; image_url: str = ""
class MaterialUpdate(BaseModel): name: Optional[str] = None; category: Optional[str] = None; category_id: Optional[int] = None; price_per_sqm: Optional[float] = None; stock_quantity: Optional[float] = None; description: Optional[str] = None; image_url: Optional[str] = None
class CategoryCreate(BaseModel): name: str; description: str = ""; image_url: str = ""
class CategoryUpdate(BaseModel): name: Optional[str] = None; description: Optional[str] = None; image_url: Optional[str] = None
class OrderItemCreate(BaseModel): material_id: str; material_name: str; width: float; height: float; quantity: int = 1; price_per_sqm: float; notes: str = ""
class OrderCreate(BaseModel): items: List[OrderItemCreate]; notes: str = ""
class OrderStatusUpdate(BaseModel): status: str; rejection_reason: str = ""
class MessageCreate(BaseModel): receiver_id: str; text: str
class AssignItemReq(BaseModel): worker_id: str
class DeliveryInfoReq(BaseModel): driver_name: str; driver_phone: str; plate_number: str = ""
class PaymentCreate(BaseModel): amount: float; note: str = ""

# ─── EXCHANGE RATE (Real-time USD/UZS) ───
@api_router.get("/exchange-rate")
async def get_exchange_rate():
    """O'zbekiston Markaziy Banki dan real vaqtda dollar kursini olish"""
    cached = cache.get("exchange_rate")
    if cached: return cached
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get("https://cbu.uz/oz/arkhiv-kursov-valyut/json/USD/")
            data = resp.json()
            if data and len(data) > 0:
                rate = float(data[0]["Rate"])
                result = {"rate": rate, "currency": "UZS", "date": data[0].get("Date", ""), "source": "CBU.uz"}
                cache.set("exchange_rate", result, 3600)
                return result
    except Exception as e:
        logger.warning(f"CBU kurs olishda xatolik: {e}")
    fallback = {"rate": 12800.0, "currency": "UZS", "date": "", "source": "fallback"}
    cache.set("exchange_rate", fallback, 300)
    return fallback

# ─── AUTH ───
@api_router.post("/auth/login")
async def login(req: LoginReq):
    db = await get_pool()
    user = await db.fetchrow("SELECT * FROM users WHERE email = $1", req.email.strip().lower())
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(401, "Email yoki parol noto'g'ri")
    token = create_access_token(str(user["id"]), user["email"], user["role"])
    u = row_to_dict(user)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    return {"token": token, "user": u}

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)): return {"user": user}

# ─── AUTH: Update own profile ───
@api_router.put("/auth/profile")
async def update_profile(request: Request, user: dict = Depends(get_current_user)):
    body = await request.json()
    new_email = body.get("email", "").strip().lower()
    new_password = body.get("password", "").strip()
    current_password = body.get("current_password", "").strip()
    if not current_password:
        raise HTTPException(400, "Joriy parolni kiriting")
    db = await get_pool()
    db_user = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(user["id"]))
    if not db_user or not verify_password(current_password, db_user["password_hash"]):
        raise HTTPException(400, "Joriy parol noto'g'ri")
    updates = []
    params = []
    param_idx = 1
    if new_email and new_email != db_user["email"]:
        existing = await db.fetchrow("SELECT id FROM users WHERE email = $1 AND id != $2", new_email, int(user["id"]))
        if existing:
            raise HTTPException(400, "Bu email allaqachon mavjud")
        updates.append(f"email = ${param_idx}")
        params.append(new_email)
        param_idx += 1
    if new_password:
        if len(new_password) < 4:
            raise HTTPException(400, "Parol kamida 4 ta belgi")
        updates.append(f"password_hash = ${param_idx}")
        params.append(hash_password(new_password))
        param_idx += 1
    if not updates:
        raise HTTPException(400, "O'zgartirish yo'q")
    params.append(int(user["id"]))
    await db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ${param_idx}", *params)
    updated = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(user["id"]))
    u = row_to_dict(updated)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    token = create_access_token(u["id"], u.get("email", ""), u.get("role", ""))
    return {"user": u, "token": token, "message": "Profil yangilandi"}

# ─── DEALERS ───
@api_router.post("/dealers")
async def create_dealer(d: DealerCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    existing = await db.fetchrow("SELECT id FROM users WHERE email = $1", d.email.strip().lower())
    if existing: raise HTTPException(400, "Email mavjud")
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,0,'',$8) RETURNING *",
        d.name, d.email.strip().lower(), hash_password(d.password), "dealer", d.phone, d.address, d.credit_limit, now
    )
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    cache.invalidate("dealers")
    return u

@api_router.get("/dealers")
async def list_dealers(admin: dict = Depends(require_admin)):
    cached = cache.get("dealers_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM users WHERE role = 'dealer' ORDER BY created_at DESC")
    out = []
    for r in rows:
        u = row_to_dict(r)
        u["id"] = str(u["id"])
        u.pop("password_hash", None)
        out.append(u)
    cache.set("dealers_list", out, 30)
    return out

@api_router.put("/dealers/{did}")
async def update_dealer(did: str, data: DealerUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []
    params = []
    idx = 1
    for field in ["name", "phone", "address", "credit_limit"]:
        val = getattr(data, field, None)
        if val is not None:
            updates.append(f"{field} = ${idx}")
            params.append(val)
            idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(int(did))
    await db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(did))
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    return u

@api_router.delete("/dealers/{did}")
async def delete_dealer(did: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM users WHERE id = $1 AND role = 'dealer'", int(did))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("dealers", "chat", "stats")
    return {"message": "Deleted"}

# ─── WORKERS ───
@api_router.post("/workers")
async def create_worker(w: WorkerCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    existing = await db.fetchrow("SELECT id FROM users WHERE email = $1", w.email.strip().lower())
    if existing: raise HTTPException(400, "Email mavjud")
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,'',$6,0,$7,$8) RETURNING *",
        w.name, w.email.strip().lower(), hash_password(w.password), "worker", w.phone, 0, w.specialty, now
    )
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    cache.invalidate("workers", "stats")
    return u

@api_router.get("/workers")
async def list_workers(admin: dict = Depends(require_admin)):
    cached = cache.get("workers_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM users WHERE role = 'worker' ORDER BY created_at DESC")
    out = []
    for r in rows:
        u = row_to_dict(r)
        u["id"] = str(u["id"])
        u.pop("password_hash", None)
        out.append(u)
    cache.set("workers_list", out, 30)
    return out

@api_router.delete("/workers/{wid}")
async def delete_worker(wid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM users WHERE id = $1 AND role = 'worker'", int(wid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("workers", "stats")
    return {"message": "Deleted"}

# ─── CATEGORIES ───
@api_router.post("/categories")
async def create_category(d: CategoryCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO categories (name, description, image_url, created_at) VALUES ($1,$2,$3,$4) RETURNING *",
        d.name, d.description, d.image_url, now
    )
    c = row_to_dict(row); c["id"] = str(c["id"])
    cache.invalidate("categories", "materials")
    return c

@api_router.get("/categories")
async def list_categories(user: dict = Depends(get_current_user)):
    cached = cache.get("categories_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM categories ORDER BY name ASC")
    out = []
    for r in rows:
        c = row_to_dict(r); c["id"] = str(c["id"])
        c["material_count"] = await db.fetchval("SELECT COUNT(*) FROM materials WHERE category_id = $1", r["id"])
        out.append(c)
    cache.set("categories_list", out, 60)
    return out

@api_router.put("/categories/{cid}")
async def update_category(cid: str, d: CategoryUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []; params = []; idx = 1
    for field in ["name", "description", "image_url"]:
        val = getattr(d, field, None)
        if val is not None:
            updates.append(f"{field} = ${idx}"); params.append(val); idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(int(cid))
    await db.execute(f"UPDATE categories SET {', '.join(updates)} WHERE id = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM categories WHERE id = $1", int(cid))
    c = row_to_dict(row); c["id"] = str(c["id"])
    cache.invalidate("categories", "materials")
    return c

@api_router.delete("/categories/{cid}")
async def delete_category(cid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    mat_count = await db.fetchval("SELECT COUNT(*) FROM materials WHERE category_id = $1", int(cid))
    if mat_count > 0:
        raise HTTPException(400, f"Bu kategoriyada {mat_count} ta mahsulot bor. Avval mahsulotlarni ko'chiring.")
    result = await db.execute("DELETE FROM categories WHERE id = $1", int(cid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("categories", "materials")
    return {"message": "Deleted"}

# ─── MATERIALS ───
@api_router.post("/materials")
async def create_material(d: MaterialCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO materials (name, category, category_id, price_per_sqm, stock_quantity, unit, description, image_url, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9) RETURNING *",
        d.name, d.category, d.category_id, d.price_per_sqm, d.stock_quantity, d.unit, d.description, d.image_url, now
    )
    m = row_to_dict(row); m["id"] = str(m["id"])
    if m.get("category_id"): m["category_id"] = str(m["category_id"])
    cache.invalidate("materials", "categories", "stats", "alerts")
    return m

@api_router.get("/materials")
async def list_materials(user: dict = Depends(get_current_user)):
    cached = cache.get("materials_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT m.*, c.name as category_name FROM materials m LEFT JOIN categories c ON m.category_id = c.id ORDER BY c.name ASC, m.name ASC")
    out = []
    for r in rows:
        m = row_to_dict(r); m["id"] = str(m["id"])
        if m.get("category_id"): m["category_id"] = str(m["category_id"])
        out.append(m)
    cache.set("materials_list", out, 60)
    return out

@api_router.get("/materials/by-category/{cid}")
async def list_materials_by_category(cid: str, user: dict = Depends(get_current_user)):
    cache_key = f"materials_cat_{cid}"
    cached = cache.get(cache_key)
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM materials WHERE category_id = $1 ORDER BY name ASC", int(cid))
    out = []
    for r in rows:
        m = row_to_dict(r); m["id"] = str(m["id"])
        if m.get("category_id"): m["category_id"] = str(m["category_id"])
        out.append(m)
    cache.set(cache_key, out, 60)
    return out

@api_router.put("/materials/{mid}")
async def update_material(mid: str, d: MaterialUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []
    params = []
    idx = 1
    for field in ["name", "category", "price_per_sqm", "stock_quantity", "description", "image_url", "category_id"]:
        val = getattr(d, field, None)
        if val is not None:
            updates.append(f"{field} = ${idx}")
            params.append(val)
            idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(int(mid))
    await db.execute(f"UPDATE materials SET {', '.join(updates)} WHERE id = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM materials WHERE id = $1", int(mid))
    m = row_to_dict(row)
    m["id"] = str(m["id"])
    cache.invalidate("materials", "categories", "alerts")
    return m

@api_router.delete("/materials/{mid}")
async def delete_material(mid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM materials WHERE id = $1", int(mid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("materials", "categories", "stats", "alerts")
    return {"message": "Deleted"}

# ─── ORDERS ───
@api_router.post("/orders")
async def create_order(data: OrderCreate, user: dict = Depends(get_current_user)):
    if user.get("role") != "dealer": raise HTTPException(403, "Faqat dilerlar")
    db = await get_pool()
    items = []; total_price = 0; total_sqm = 0
    for it in data.items:
        sqm = it.width * it.height * it.quantity
        price = sqm * it.price_per_sqm
        total_sqm += sqm; total_price += price
        items.append({"material_id": it.material_id, "material_name": it.material_name, "width": it.width, "height": it.height, "quantity": it.quantity, "sqm": round(sqm, 2), "price_per_sqm": it.price_per_sqm, "price": round(price, 2), "notes": it.notes, "assigned_worker_id": "", "assigned_worker_name": "", "worker_status": "pending"})
    order_code = generate_order_code()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO orders (order_code, dealer_id, dealer_name, items, total_sqm, total_price, status, notes, rejection_reason, delivery_info, created_at, updated_at) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12) RETURNING *",
        order_code, int(user["id"]), user.get("name", ""), json.dumps(items), round(total_sqm, 2), round(total_price, 2), "kutilmoqda", data.notes, "", None, now, now
    )
    o = row_to_dict(row)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    await db.execute("UPDATE users SET debt = debt + $1 WHERE id = $2", total_price, int(user["id"]))
    cache.invalidate("orders", "stats", "reports")
    return o

@api_router.get("/orders")
async def list_orders(user: dict = Depends(get_current_user)):
    cache_key = f"orders_{user['id']}_{user.get('role','')}"
    cached = cache.get(cache_key)
    if cached: return cached
    db = await get_pool()
    if user.get("role") == "dealer":
        rows = await db.fetch("SELECT * FROM orders WHERE dealer_id = $1 ORDER BY created_at DESC", int(user["id"]))
    else:
        rows = await db.fetch("SELECT * FROM orders ORDER BY created_at DESC")
    out = []
    for r in rows:
        o = row_to_dict(r)
        o["id"] = str(o["id"])
        o["dealer_id"] = str(o["dealer_id"])
        o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
        o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
        out.append(o)
    cache.set(cache_key, out, 15)
    return out

@api_router.get("/orders/{oid}")
async def get_order(oid: str, user: dict = Depends(get_current_user)):
    db = await get_pool()
    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    if not o: raise HTTPException(404, "Not found")
    o = row_to_dict(o)
    if user.get("role") == "dealer" and str(o["dealer_id"]) != user["id"]: raise HTTPException(403)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    return o

@api_router.put("/orders/{oid}/status")
async def update_order_status(oid: str, data: OrderStatusUpdate, admin: dict = Depends(require_admin)):
    valid = ["kutilmoqda","tasdiqlangan","tayyorlanmoqda","tayyor","yetkazilmoqda","yetkazildi","rad_etilgan"]
    if data.status not in valid: raise HTTPException(400, "Invalid status")
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    if data.status == "rad_etilgan" and data.rejection_reason:
        await db.execute("UPDATE orders SET status = $1, rejection_reason = $2, updated_at = $3 WHERE id = $4", data.status, data.rejection_reason, now, int(oid))
    else:
        await db.execute("UPDATE orders SET status = $1, updated_at = $2 WHERE id = $3", data.status, now, int(oid))
    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    o = row_to_dict(o)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    cache.invalidate("orders", "stats", "reports")
    return o

# ─── WORKER: Assign item to worker ───
@api_router.put("/orders/{oid}/items/{item_idx}/assign")
async def assign_item_to_worker(oid: str, item_idx: int, data: AssignItemReq, admin: dict = Depends(require_admin)):
    db = await get_pool()
    order = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    if not order: raise HTTPException(404, "Order not found")
    items = json.loads(order["items"]) if isinstance(order["items"], str) else order["items"]
    if item_idx >= len(items): raise HTTPException(400, "Invalid item index")
    worker = await db.fetchrow("SELECT * FROM users WHERE id = $1 AND role = 'worker'", int(data.worker_id))
    if not worker: raise HTTPException(404, "Worker not found")
    items[item_idx]["assigned_worker_id"] = data.worker_id
    items[item_idx]["assigned_worker_name"] = worker["name"]
    items[item_idx]["worker_status"] = "assigned"
    await db.execute("UPDATE orders SET items = $1 WHERE id = $2", json.dumps(items), int(oid))
    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    o = row_to_dict(o)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    cache.invalidate("orders")
    return o

# ─── WORKER: Get my assigned items ───
@api_router.get("/worker/tasks")
async def get_worker_tasks(user: dict = Depends(get_current_user)):
    if user.get("role") != "worker": raise HTTPException(403)
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM orders WHERE status IN ('tasdiqlangan','tayyorlanmoqda')")
    tasks = []
    for r in rows:
        o = row_to_dict(r)
        items = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
        for idx, item in enumerate(items):
            if item.get("assigned_worker_id") == user["id"]:
                tasks.append({"order_id": str(o["id"]), "order_code": o.get("order_code", ""), "dealer_name": o.get("dealer_name", ""), "item_index": idx, "material_name": item["material_name"], "width": item["width"], "height": item["height"], "sqm": item["sqm"], "notes": item.get("notes", ""), "worker_status": item.get("worker_status", "assigned"), "created_at": o["created_at"]})
    return tasks

# ─── WORKER: Mark item as completed ───
@api_router.put("/worker/tasks/{oid}/{item_idx}/complete")
async def complete_worker_task(oid: str, item_idx: int, user: dict = Depends(get_current_user)):
    if user.get("role") != "worker": raise HTTPException(403)
    db = await get_pool()
    order = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    if not order: raise HTTPException(404)
    items = json.loads(order["items"]) if isinstance(order["items"], str) else order["items"]
    if item_idx >= len(items): raise HTTPException(400)
    if items[item_idx].get("assigned_worker_id") != user["id"]: raise HTTPException(403, "Not your task")
    if items[item_idx].get("worker_status") == "completed":
        # Already completed - return current order without error
        o = row_to_dict(order)
        o["id"] = str(o["id"]); o["dealer_id"] = str(o["dealer_id"])
        o["items"] = items
        o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
        return o
    items[item_idx]["worker_status"] = "completed"
    now = datetime.now(timezone.utc).isoformat()
    await db.execute("UPDATE orders SET items = $1, updated_at = $2 WHERE id = $3", json.dumps(items), now, int(oid))

    # Check if all assigned items are completed
    all_done = all(it.get("worker_status") == "completed" for it in items if it.get("assigned_worker_id"))
    if all_done:
        await db.execute("UPDATE orders SET status = 'tayyor', updated_at = $1 WHERE id = $2", now, int(oid))
        # Send auto-message to dealer
        admin = await db.fetchrow("SELECT id, name FROM users WHERE role = 'admin' LIMIT 1")
        if admin and order["dealer_id"]:
            await db.execute(
                "INSERT INTO messages (sender_id, sender_name, sender_role, receiver_id, text, read, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7)",
                admin["id"], admin["name"] or "Admin", "admin", order["dealer_id"],
                f"Buyurtma #{order['order_code']} tayyor! Barcha ishlar tugallandi.",
                False, now
            )
        logger.info(f"Buyurtma #{order['order_code']} tayyor — dilerga xabar yuborildi")
    else:
        # Single item completed — notify admin via message
        admin = await db.fetchrow("SELECT id FROM users WHERE role = 'admin' LIMIT 1")
        if admin:
            completed_count = sum(1 for it in items if it.get("worker_status") == "completed")
            total_assigned = sum(1 for it in items if it.get("assigned_worker_id"))
            await db.execute(
                "INSERT INTO messages (sender_id, sender_name, sender_role, receiver_id, text, read, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7)",
                int(user["id"]), user.get("name", "Ishchi"), "worker", admin["id"],
                f"#{order['order_code']}: {items[item_idx]['material_name']} tayyor ({completed_count}/{total_assigned})",
                False, now
            )

    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    o = row_to_dict(o)
    o["id"] = str(o["id"]); o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    cache.invalidate("orders", "stats", "reports")
    return o

# ─── DELIVERY: Assign delivery info directly to order ───
@api_router.put("/orders/{oid}/delivery")
async def assign_delivery(oid: str, data: DeliveryInfoReq, admin: dict = Depends(require_admin)):
    db = await get_pool()
    order = await db.fetchrow("SELECT id FROM orders WHERE id = $1", int(oid))
    if not order: raise HTTPException(404, "Buyurtma topilmadi")
    d_info = json.dumps({"driver_name": data.driver_name, "driver_phone": data.driver_phone, "plate_number": data.plate_number})
    now = datetime.now(timezone.utc).isoformat()
    await db.execute("UPDATE orders SET delivery_info = $1, status = 'yetkazilmoqda', updated_at = $2 WHERE id = $3", d_info, now, int(oid))
    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    o = row_to_dict(o)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    cache.invalidate("orders", "stats")
    return o

# ─── DELIVERY: Admin confirms delivery ───
@api_router.put("/orders/{oid}/confirm-delivery")
async def confirm_delivery(oid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    order = await db.fetchrow("SELECT id FROM orders WHERE id = $1", int(oid))
    if not order: raise HTTPException(404, "Buyurtma topilmadi")
    now = datetime.now(timezone.utc).isoformat()
    await db.execute("UPDATE orders SET status = 'yetkazildi', updated_at = $1 WHERE id = $2", now, int(oid))
    o = await db.fetchrow("SELECT * FROM orders WHERE id = $1", int(oid))
    o = row_to_dict(o)
    o["id"] = str(o["id"])
    o["dealer_id"] = str(o["dealer_id"])
    o["items"] = json.loads(o["items"]) if isinstance(o["items"], str) else o["items"]
    o["delivery_info"] = json.loads(o["delivery_info"]) if isinstance(o["delivery_info"], str) and o["delivery_info"] else o["delivery_info"]
    cache.invalidate("orders", "stats", "reports")
    return o

# ─── CHAT ───
@api_router.post("/messages")
async def send_message(data: MessageCreate, user: dict = Depends(get_current_user)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO messages (sender_id, sender_name, sender_role, receiver_id, text, read, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7) RETURNING *",
        int(user["id"]), user.get("name", ""), user.get("role", ""), int(data.receiver_id), data.text, False, now
    )
    m = row_to_dict(row)
    m["id"] = str(m["id"])
    m["sender_id"] = str(m["sender_id"])
    m["receiver_id"] = str(m["receiver_id"])
    cache.invalidate("chat")
    return m

@api_router.get("/messages/{pid}")
async def get_messages(pid: str, user: dict = Depends(get_current_user)):
    db = await get_pool()
    rows = await db.fetch(
        "SELECT * FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at ASC",
        int(user["id"]), int(pid)
    )
    out = []
    for r in rows:
        m = row_to_dict(r)
        m["id"] = str(m["id"])
        m["sender_id"] = str(m["sender_id"])
        m["receiver_id"] = str(m["receiver_id"])
        out.append(m)
    await db.execute("UPDATE messages SET read = TRUE WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE", int(pid), int(user["id"]))
    return out

@api_router.get("/chat/partners")
async def get_chat_partners(user: dict = Depends(get_current_user)):
    db = await get_pool()
    if user.get("role") == "admin":
        rows = await db.fetch("SELECT * FROM users WHERE role = 'dealer' ORDER BY name")
        out = []
        for r in rows:
            d = row_to_dict(r)
            d["id"] = str(d["id"])
            d.pop("password_hash", None)
            lm = await db.fetchrow(
                "SELECT text, created_at FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at DESC LIMIT 1",
                int(user["id"]), r["id"]
            )
            uc = await db.fetchval(
                "SELECT COUNT(*) FROM messages WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE",
                r["id"], int(user["id"])
            )
            d["last_message"] = lm["text"] if lm else ""
            d["last_message_time"] = lm["created_at"] if lm else ""
            d["unread_count"] = uc or 0
            out.append(d)
        return out
    else:
        admin = await db.fetchrow("SELECT * FROM users WHERE role = 'admin' LIMIT 1")
        if not admin: return []
        a = row_to_dict(admin)
        a["id"] = str(a["id"])
        a.pop("password_hash", None)
        lm = await db.fetchrow(
            "SELECT text, created_at FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at DESC LIMIT 1",
            int(user["id"]), admin["id"]
        )
        uc = await db.fetchval(
            "SELECT COUNT(*) FROM messages WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE",
            admin["id"], int(user["id"])
        )
        a["last_message"] = lm["text"] if lm else ""
        a["last_message_time"] = lm["created_at"] if lm else ""
        a["unread_count"] = uc or 0
        return [a]

# ─── IMAGE UPLOAD ───
@api_router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Faqat rasm fayllari ruxsat etiladi")
    ext = file.filename.split(".")[-1] if file.filename and "." in file.filename else "jpg"
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = UPLOAD_DIR / filename
    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)
    image_url = f"/api/uploads/{filename}"
    return {"image_url": image_url}

# ─── STATISTICS ───
@api_router.get("/statistics")
async def get_statistics(admin: dict = Depends(require_admin)):
    cached = cache.get("stats_all")
    if cached: return cached
    db = await get_pool()
    total_revenue = await db.fetchval("SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE status IN ('tasdiqlangan','tayyorlanmoqda','tayyor','yetkazilmoqda','yetkazildi')")
    result = {
        "total_orders": await db.fetchval("SELECT COUNT(*) FROM orders"),
        "pending_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'kutilmoqda'"),
        "approved_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'tasdiqlangan'"),
        "preparing_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'tayyorlanmoqda'"),
        "ready_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'tayyor'"),
        "delivering_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'yetkazilmoqda'"),
        "delivered_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'yetkazildi'"),
        "rejected_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status = 'rad_etilgan'"),
        "total_dealers": await db.fetchval("SELECT COUNT(*) FROM users WHERE role = 'dealer'"),
        "total_workers": await db.fetchval("SELECT COUNT(*) FROM users WHERE role = 'worker'"),
        "total_materials": await db.fetchval("SELECT COUNT(*) FROM materials"),
        "total_revenue": round(float(total_revenue), 2),
    }
    cache.set("stats_all", result, 30)
    return result

# ─── SEED & STARTUP ───
async def create_tables(db):
    await db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'dealer',
            phone TEXT DEFAULT '',
            address TEXT DEFAULT '',
            credit_limit FLOAT DEFAULT 0,
            debt FLOAT DEFAULT 0,
            specialty TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            image_url TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS materials (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT '',
            category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL,
            price_per_sqm FLOAT NOT NULL DEFAULT 0,
            stock_quantity FLOAT NOT NULL DEFAULT 0,
            unit TEXT DEFAULT 'kv.m',
            description TEXT DEFAULT '',
            image_url TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    # Add category_id column if not exists (migration)
    try:
        await db.execute("ALTER TABLE materials ADD COLUMN IF NOT EXISTS category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL")
    except Exception:
        pass
    await db.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id SERIAL PRIMARY KEY,
            order_code TEXT DEFAULT '',
            dealer_id INTEGER REFERENCES users(id),
            dealer_name TEXT DEFAULT '',
            items TEXT DEFAULT '[]',
            total_sqm FLOAT DEFAULT 0,
            total_price FLOAT DEFAULT 0,
            status TEXT DEFAULT 'kutilmoqda',
            notes TEXT DEFAULT '',
            rejection_reason TEXT DEFAULT '',
            delivery_info TEXT,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id SERIAL PRIMARY KEY,
            sender_id INTEGER REFERENCES users(id),
            sender_name TEXT DEFAULT '',
            sender_role TEXT DEFAULT '',
            receiver_id INTEGER REFERENCES users(id),
            text TEXT DEFAULT '',
            read BOOLEAN DEFAULT FALSE,
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id SERIAL PRIMARY KEY,
            dealer_id INTEGER REFERENCES users(id),
            amount FLOAT NOT NULL DEFAULT 0,
            note TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)

async def seed_admin(db):
    email = os.environ.get("ADMIN_EMAIL", "admin@curtain.uz")
    pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    now = datetime.now(timezone.utc).isoformat()
    ex = await db.fetchrow("SELECT * FROM users WHERE email = $1", email)
    if not ex:
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,'','',0,0,'',$5)",
            "Admin", email, hash_password(pw), "admin", now
        )
        logger.info(f"Admin yaratildi: {email}")
    elif not verify_password(pw, ex["password_hash"]):
        await db.execute("UPDATE users SET password_hash = $1 WHERE email = $2", hash_password(pw), email)

    mat_count = await db.fetchval("SELECT COUNT(*) FROM materials")
    if mat_count == 0:
        # Create default categories
        cat_count = await db.fetchval("SELECT COUNT(*) FROM categories")
        if cat_count == 0:
            cats = [
                ("Parda", "Har xil parda turlari"),
                ("Jalyuzi", "Gorizontal va vertikal jalyuzilar"),
                ("Aksessuar", "Karniz, gardina va boshqa aksessuarlar"),
            ]
            for c in cats:
                await db.execute("INSERT INTO categories (name, description, created_at) VALUES ($1,$2,$3)", c[0], c[1], now)
            logger.info("Kategoriyalar yaratildi")

        parda_id = await db.fetchval("SELECT id FROM categories WHERE name = 'Parda'")
        jalyuzi_id = await db.fetchval("SELECT id FROM categories WHERE name = 'Jalyuzi'")

        materials_data = [
            ("Blackout Parda", "Parda", parda_id, 7.0, 500, "kv.m", "Yorug'lik o'tkazmaydigan parda", "https://images.pexels.com/photos/4814070/pexels-photo-4814070.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Tull Parda", "Parda", parda_id, 3.5, 800, "kv.m", "Shaffof tull parda", "https://images.unsplash.com/photo-1574197635162-68e4b468e4e9?w=600"),
            ("Roller Jalyuzi", "Jalyuzi", jalyuzi_id, 10.0, 300, "kv.m", "Zamonaviy roller jalyuzi", "https://images.pexels.com/photos/19166538/pexels-photo-19166538.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Gorizontal Jalyuzi", "Jalyuzi", jalyuzi_id, 8.0, 400, "kv.m", "Alyuminiy gorizontal jalyuzi", "https://images.unsplash.com/photo-1603299938527-d035bc6fc2c8?w=600"),
            ("Vertikal Jalyuzi", "Jalyuzi", jalyuzi_id, 6.0, 350, "kv.m", "Ofis uchun vertikal jalyuzi", "https://images.pexels.com/photos/8955198/pexels-photo-8955198.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Rimskaya Parda", "Parda", parda_id, 9.0, 200, "kv.m", "Premium rimskaya parda", "https://images.unsplash.com/photo-1729277980958-092c5e9e2ea4?w=600"),
        ]
        for m in materials_data:
            await db.execute(
                "INSERT INTO materials (name, category, category_id, price_per_sqm, stock_quantity, unit, description, image_url, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)",
                m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], now
            )
        logger.info("Materiallar yaratildi")

    if not await db.fetchrow("SELECT id FROM users WHERE email = 'dealer@test.uz'"):
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,0,'',$8)",
            "Test Diler", "dealer@test.uz", hash_password("dealer123"), "dealer", "+998901234567", "Toshkent, Yunusobod", 5000.0, now
        )
        logger.info("Demo diler yaratildi")

    if not await db.fetchrow("SELECT id FROM users WHERE email = 'worker@test.uz'"):
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,'',$6,0,$7,$8)",
            "Aziz Ishchi", "worker@test.uz", hash_password("worker123"), "worker", "+998901112233", 0.0, "Jalyuzi o'rnatish", now
        )
        logger.info("Demo ishchi yaratildi")

# ─── REPORTS - Hisobot tizimi ───
@api_router.get("/reports")
async def get_reports(admin: dict = Depends(require_admin)):
    cached = cache.get("reports_all")
    if cached: return cached
    db = await get_pool()
    now = datetime.now(timezone.utc)

    # Weekly & Monthly revenue
    week_ago = (now - timedelta(days=7)).isoformat()
    month_ago = (now - timedelta(days=30)).isoformat()

    weekly_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND status NOT IN ('rad_etilgan')", week_ago
    )
    monthly_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND status NOT IN ('rad_etilgan')", month_ago
    )
    total_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE status NOT IN ('rad_etilgan')"
    )
    weekly_orders = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1", week_ago)
    monthly_orders = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1", month_ago)
    total_orders = await db.fetchval("SELECT COUNT(*) FROM orders")

    # Top selling materials (from order items)
    all_orders = await db.fetch("SELECT items FROM orders WHERE status NOT IN ('rad_etilgan')")
    mat_stats: dict = {}
    for row in all_orders:
        items = json.loads(row["items"]) if isinstance(row["items"], str) else row["items"]
        for it in items:
            name = it.get("material_name", "Noma'lum")
            sqm = it.get("sqm", 0)
            price = it.get("price", 0)
            if name not in mat_stats:
                mat_stats[name] = {"name": name, "total_sqm": 0, "total_price": 0, "count": 0}
            mat_stats[name]["total_sqm"] += sqm
            mat_stats[name]["total_price"] += price
            mat_stats[name]["count"] += 1

    top_materials = sorted(mat_stats.values(), key=lambda x: x["total_price"], reverse=True)[:5]

    # Top dealers
    dealer_rows = await db.fetch("""
        SELECT u.name, COUNT(o.id) as order_count, COALESCE(SUM(o.total_price), 0) as revenue
        FROM orders o JOIN users u ON o.dealer_id = u.id
        WHERE o.status NOT IN ('rad_etilgan')
        GROUP BY u.name ORDER BY revenue DESC LIMIT 5
    """)
    top_dealers = [{"name": r["name"], "orders": r["order_count"], "revenue": round(float(r["revenue"]), 2)} for r in dealer_rows]

    # Daily orders for last 7 days
    daily = []
    for i in range(6, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0).isoformat()
        day_end = (now - timedelta(days=i)).replace(hour=23, minute=59, second=59).isoformat()
        cnt = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1 AND created_at <= $2", day_start, day_end)
        rev = await db.fetchval("SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND created_at <= $2 AND status NOT IN ('rad_etilgan')", day_start, day_end)
        day_label = (now - timedelta(days=i)).strftime("%d.%m")
        daily.append({"day": day_label, "orders": cnt, "revenue": round(float(rev), 2)})

    result = {
        "weekly_revenue": round(float(weekly_revenue), 2),
        "monthly_revenue": round(float(monthly_revenue), 2),
        "total_revenue": round(float(total_revenue), 2),
        "weekly_orders": weekly_orders,
        "monthly_orders": monthly_orders,
        "total_orders": total_orders,
        "top_materials": top_materials,
        "top_dealers": top_dealers,
        "daily": daily,
    }
    cache.set("reports_all", result, 60)
    return result

# ─── LOW STOCK ALERTS ───
@api_router.get("/alerts/low-stock")
async def get_low_stock(admin: dict = Depends(require_admin)):
    cached = cache.get("alerts_low_stock")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM materials WHERE stock_quantity < 10 ORDER BY stock_quantity ASC")
    out = []
    for r in rows:
        m = row_to_dict(r)
        m["id"] = str(m["id"])
        out.append(m)
    cache.set("alerts_low_stock", out, 60)
    return out

# ─── EXCEL EXPORT ───
@api_router.get("/reports/export-orders")
async def export_orders_excel(admin: dict = Depends(require_admin)):
    db = await get_pool()
    orders = await db.fetch("SELECT * FROM orders ORDER BY created_at DESC")

    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    # Styling
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ['#', 'Buyurtma kodi', 'Diler', 'Mahsulotlar', 'Jami kv.m', 'Jami narx ($)', 'Status', 'Sana']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for idx, order in enumerate(orders, 1):
        items = json.loads(order["items"]) if isinstance(order["items"], str) else order["items"]
        item_names = ", ".join([f'{it["material_name"]} ({it.get("width", 0)}x{it.get("height", 0)}m)' for it in items])
        status_map = {"kutilmoqda": "Kutilmoqda", "tasdiqlangan": "Tasdiqlangan", "tayyorlanmoqda": "Tayyorlanmoqda", "tayyor": "Tayyor", "yetkazilmoqda": "Yetkazilmoqda", "yetkazildi": "Yetkazildi", "rad_etilgan": "Rad etilgan"}

        row = [idx, order["order_code"], order["dealer_name"], item_names, round(order["total_sqm"], 2), round(order["total_price"], 2), status_map.get(order["status"], order["status"]), order["created_at"][:16].replace("T", " ")]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=col, value=val)
            cell.border = border
            if col in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = datetime.now(timezone.utc).strftime('%Y%m%d')

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=buyurtmalar_{today}.xlsx"}
    )

# ─── DEALER PAYMENTS (To'lovlar) ───
@api_router.post("/dealers/{did}/payment")
async def add_dealer_payment(did: str, data: PaymentCreate, admin: dict = Depends(require_admin)):
    if data.amount <= 0: raise HTTPException(400, "Summa 0 dan katta bo'lishi kerak")
    db = await get_pool()
    dealer = await db.fetchrow("SELECT * FROM users WHERE id = $1 AND role = 'dealer'", int(did))
    if not dealer: raise HTTPException(404, "Diler topilmadi")
    now = datetime.now(timezone.utc).isoformat()
    await db.execute(
        "INSERT INTO payments (dealer_id, amount, note, created_at) VALUES ($1,$2,$3,$4)",
        int(did), data.amount, data.note, now
    )
    new_debt = max(0, (dealer["debt"] or 0) - data.amount)
    await db.execute("UPDATE users SET debt = $1 WHERE id = $2", new_debt, int(did))
    cache.invalidate("dealers", "stats")
    return {"message": "To'lov qabul qilindi", "new_debt": round(new_debt, 2), "paid": data.amount}

@api_router.get("/dealers/{did}/payments")
async def get_dealer_payments(did: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM payments WHERE dealer_id = $1 ORDER BY created_at DESC", int(did))
    out = []
    for r in rows:
        p = row_to_dict(r)
        p["id"] = str(p["id"])
        p["dealer_id"] = str(p["dealer_id"])
        out.append(p)
    return out

# ─── HEALTH CHECK ───
@api_router.get("/health")
async def health_check():
    try:
        db = await get_pool()
        await db.fetchval("SELECT 1")
        return {"status": "ok", "database": "connected", "time": datetime.now(timezone.utc).isoformat()}
    except Exception as e:
        return {"status": "error", "database": str(e)}

# ─── KEEP ALIVE - PostgreSQL uxlab qolmasligi uchun ───
async def keep_alive_task():
    """Har 5 daqiqada PostgreSQL'ga ping yuborib, uxlab qolmasligini ta'minlaydi"""
    while True:
        try:
            await asyncio.sleep(300)  # 5 daqiqa
            db = await get_pool()
            await db.fetchval("SELECT 1")
            logger.info("🟢 Keep-alive ping → PostgreSQL OK")
        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.warning(f"🔴 Keep-alive ping xatolik: {e}")

@app.on_event("startup")
async def startup():
    global pool
    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)
    async with pool.acquire() as conn:
        await create_tables(conn)
        await seed_admin(conn)
    asyncio.create_task(keep_alive_task())
    logger.info("Server ishga tushdi! (PostgreSQL + Keep-Alive)")

@app.on_event("shutdown")
async def shutdown():
    global pool
    if pool:
        await pool.close()

app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
